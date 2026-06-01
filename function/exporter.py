import csv

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import (
    EXCEL_COL_WIDTHS,
    EXCEL_HEADER_COLOR,
    EXCEL_ROW_EVEN,
    EXCEL_ROW_ODD,
    OUTPUT_COLUMNS,
)


def save_csv(data: list[dict], output_path: str) -> None:
    """Ghi danh sách record ra file CSV (UTF-8 BOM để Excel mở đúng tiếng Việt)."""
    if not data:
        return
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)


def save_excel(data: list[dict], output_path: str) -> None:
    """Ghi danh sách record ra file Excel với định dạng header + alternating rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # ── Style definitions ────────────────────────────────────
    thin_side   = Side(style="thin", color="BFBFBF")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    header_fill = PatternFill("solid", start_color=EXCEL_HEADER_COLOR, end_color=EXCEL_HEADER_COLOR)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    row_font         = Font(name="Arial", size=10)
    row_align_left   = Alignment(horizontal="left",   vertical="center")
    row_align_center = Alignment(horizontal="center", vertical="center")
    fill_even = PatternFill("solid", start_color=EXCEL_ROW_EVEN, end_color=EXCEL_ROW_EVEN)
    fill_odd  = PatternFill("solid", start_color=EXCEL_ROW_ODD,  end_color=EXCEL_ROW_ODD)

    # ── Header row ───────────────────────────────────────────
    for col_idx, (header, width) in enumerate(zip(OUTPUT_COLUMNS, EXCEL_COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22

    # ── Data rows ────────────────────────────────────────────
    for row_idx, record in enumerate(data, start=2):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.fill      = fill
            cell.font      = row_font
            cell.border    = cell_border
            cell.alignment = row_align_left if col_idx == 1 else row_align_center
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"
    wb.save(output_path)